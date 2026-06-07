"""Spreadsheet round-trip: export an editable .xlsx, import it back.

This is the manual-entry path for a NON-TECHNICAL editor (city staff). It turns
the database into a friendly Excel workbook they can fill by hand, then reads
that workbook back into the pipeline.

The workbook is a **per-agency calendar-year time-series**: one tab per agency,
plus two reference tabs.

  1. "How to use"      -- plain-language instructions and the colour legend.
  2. "Data Dictionary" -- one row per metric (all 32): what each means, its unit,
                          formula, native frequency.
  3..N  one tab PER AGENCY (e.g. "TTC", "STM", ...). Each agency tab is a grid:
        one ROW per metric, and COLUMNS grouped by CALENDAR year, each year a
        block of:  M1 M2 M3 Q1 | M4 M5 M6 Q2 | M7 M8 M9 Q3 | M10 M11 M12 Q4 | YTD | Year
        The editor TYPES the raw monthly cells they have (white). The Q / YTD /
        Year cells are READ-ONLY computed cells -- live Excel SUM formulas over
        the month cells (presentation only; the authoritative roll-up is the
        server-side rollup job, never the workbook). Annual-only metrics are
        typed once, into the Year cell; their month/quarter cells are locked
        (grey). Derived ratios show as a grey live Excel formula in the Year cell.

        Each agency tab also carries a per-mode FLEET block: one white column per
        weighted mode (Bus / Subway / Light rail / Commuter rail / Streetcar) that
        IMPORTS into metric `fleet_size` at that mode, plus a grey computed
        "Fleet scale" (fleet_capacity) cell.

Colours: white = type here · grey = calculated / do-not-touch. Every grey cell is
recomputed on the server, so import reads ONLY the white cells (never a grey /
derived / roll-up cell) and never fabricates a blank.

Display names + plain meanings come from the per-metric data dictionary
(`dictionary.load_dictionary()`). openpyxl AND PyYAML are imported LAZILY inside
export/import, so merely importing this module never requires either.
"""

from __future__ import annotations

from .equations import RatioEquation, defining_equation
from .refdata import METRICS, MODE_CAPACITY_WEIGHT, NON_RANKABLE_METRICS

# --- Metric groupings, agency + mode names -----------------------------------

# The two monthly-native feeds (StatCan 23-10-0307 publishes both): typed
# month-by-month; their Q/YTD/Year cells are live SUM roll-ups.
MONTHLY_METRICS: list[str] = ["ridership", "operating_revenue"]

# fleet_size is entered PER MODE in the Fleet block, not as a system-wide row;
# fleet_capacity is the derived Fleet-scale cell. Both are kept out of the main
# metric rows.
_FLEET_SIZE = "fleet_size"
_FLEET_CAPACITY = "fleet_capacity"

# System-wide sourced metrics typed once a year (everything sourced except the two
# monthly feeds and the per-mode fleet_size). In METRICS order.
ANNUAL_SOURCED_METRICS: list[str] = [
    c for c, m in METRICS.items()
    if not m["is_derived"] and c not in MONTHLY_METRICS and c not in (_FLEET_SIZE, _FLEET_CAPACITY)
]
# Derived system-wide metrics shown as a live Year-cell formula. fleet_capacity is
# excluded here -- it lives in the Fleet block as its own Fleet-scale cell.
ANNUAL_DERIVED_METRICS: list[str] = [
    c for c, m in METRICS.items() if m["is_derived"] and c != _FLEET_CAPACITY
]

# The full ordered list of metric ROWS on an agency tab: monthly feeds first, then
# annual sourced, then derived. fleet_size / fleet_capacity are NOT here (Fleet block).
METRIC_ROWS: list[str] = MONTHLY_METRICS + ANNUAL_SOURCED_METRICS + ANNUAL_DERIVED_METRICS

# The weighted modes whose fleet_size the user enters, in display order (the modes
# that carry a capacity weight; Σ weight × fleet_size = fleet_capacity).
FLEET_MODES: list[tuple[str, str]] = [
    ("bus", "Bus"),
    ("subway", "Subway"),
    ("light_rail", "Light rail"),
    ("commuter_rail", "Commuter rail"),
    ("streetcar", "Streetcar"),
]

# Agency slug -> short name (= tab title). Reverse-lookup map on import.
AGENCY_NAMES: dict[str, str] = {
    "ttc": "TTC",
    "stm": "STM",
    "translink": "TransLink",
    "metrolinx": "Metrolinx",
    "oc-transpo": "OC Transpo",
    "calgary-transit": "Calgary Transit",
    "edmonton-ets": "ETS",
    "miway": "MiWay",
    "bc-transit": "BC Transit",
    "burlington-transit": "Burlington Transit",
    # expansion agencies
    "winnipeg-transit": "Winnipeg Transit",
    "hamilton-street-railway": "HSR",
    "brampton-transit": "Brampton Transit",
    "grand-river-transit": "GRT",
    "stl-laval": "STL",
    "rtl-longueuil": "RTL",
    "york-region-transit": "YRT",
    "halifax-transit": "Halifax Transit",
    "durham-region-transit": "DRT",
    "saskatoon-transit": "Saskatoon Transit",
    "regina-transit": "Regina Transit",
}

# Reference sheet names.
SHEET_HOWTO = "How to use"
SHEET_DICT = "Data Dictionary"

# Within-year column geometry. 18 columns per calendar year:
#   M1 M2 M3 Q1 | M4 M5 M6 Q2 | M7 M8 M9 Q3 | M10 M11 M12 Q4 | YTD | Year
_MONTH_ABBR = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
# Sub-header label + a parallel "kind" tag for each of the 18 columns, used by both
# export (styling/formulas) and import (which cells to read).
_YEAR_SUBHEADERS: list[tuple[str, str]] = []
for _q in range(4):
    for _mi in range(3):
        _YEAR_SUBHEADERS.append((_MONTH_ABBR[_q * 3 + _mi], "month"))
    _YEAR_SUBHEADERS.append((f"Q{_q + 1}", "quarter"))
_YEAR_SUBHEADERS.append(("YTD", "ytd"))
_YEAR_SUBHEADERS.append(("Year", "year"))
YEAR_BLOCK_WIDTH = len(_YEAR_SUBHEADERS)  # 18

# 0-based offsets within a year block.
_MONTH_OFFSETS = [i for i, (_l, k) in enumerate(_YEAR_SUBHEADERS) if k == "month"]  # 12
_QUARTER_OFFSETS = [i for i, (_l, k) in enumerate(_YEAR_SUBHEADERS) if k == "quarter"]
_YTD_OFFSET = next(i for i, (_l, k) in enumerate(_YEAR_SUBHEADERS) if k == "ytd")
_YEAR_OFFSET = next(i for i, (_l, k) in enumerate(_YEAR_SUBHEADERS) if k == "year")

# Layout anchors (1-based). Column A is the row label; year blocks start at col B.
_LABEL_COL = 1
_FIRST_YEAR_COL = 2
# Rows: 1 = agency title, 2 = year-block header, 3 = within-year sub-header, then data.
_TITLE_ROW = 1
_YEAR_HEADER_ROW = 2
_SUBHEADER_ROW = 3
_FIRST_DATA_ROW = 4

# Fills (ARGB). White = no fill.
_GREY = "FFD9D9D9"


def _native_frequency(code: str) -> str:
    return "Monthly" if code in MONTHLY_METRICS else "Annual"


def _excel_unit_format(code: str) -> str:
    """openpyxl number_format string chosen from the metric's unit/unit_type."""
    meta = METRICS[code]
    unit, unit_type = meta["unit"], meta["unit_type"]
    if unit_type == "currency":
        return '#,##0.00'
    if unit == "%":
        return '#,##0.0'
    if unit_type == "count":
        return '#,##0'
    return '#,##0.00'


def _col(idx: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(idx)


def _year_start_col(year_idx: int) -> int:
    """1-based column where calendar-year `year_idx`'s 18-col block begins."""
    return _FIRST_YEAR_COL + year_idx * YEAR_BLOCK_WIDTH


# --- Formula plumbing (the Q/YTD/Year + Fleet-scale read-only cells) ----------


def _sum_formula(row: int, cols: list[int]) -> str:
    """A SUM over the given 1-based columns in `row`; blank when all are empty."""
    refs = [f"{_col(c)}{row}" for c in cols]
    joined = ",".join(refs)
    # COUNT==0 -> blank, so an untouched roll-up shows nothing (not a zero).
    return f'=IF(COUNT({joined})=0,"",SUM({joined}))'


def _derived_year_formula(code: str, year_start: int, year_rows: dict[str, int]) -> str:
    """Live Excel formula for a derived metric's Year cell, mirroring its equation.

    Operands are read from each operand metric's own Year column in the same year
    block (`year_rows`: metric code -> its data row). Blank when an input cell is
    empty or a ratio denominator is zero. An operand with no row on this tab (e.g.
    an `attr:` agency attribute, or a metric defined by a non-SUM/RATIO rule) ->
    blank cell; the server is the source of truth there.
    """
    eq = defining_equation(code)
    if eq is None:
        return ""
    if isinstance(eq, RatioEquation):
        needed = [eq.numerator, eq.denominator]
    else:
        needed = [t for _s, t in eq.terms]
    if any(c not in year_rows for c in needed):
        return ""

    def ref(c: str) -> str:
        return f"{_col(year_start + _YEAR_OFFSET)}{year_rows[c]}"

    if isinstance(eq, RatioEquation):
        num, den = ref(eq.numerator), ref(eq.denominator)
        return f'=IF(OR({num}="",{den}="",{den}=0),"",{num}/{den})'

    cells = [ref(t) for _s, t in eq.terms]
    guard = "OR(" + ",".join(f'{c}=""' for c in cells) + ")"
    parts: list[str] = []
    for i, (sign, term) in enumerate(eq.terms):
        op = "-" if sign < 0 else ("+" if i else "")
        parts.append(f"{op}{ref(term)}")
    return f'=IF({guard},"",{"".join(parts)})'


def _fleet_capacity_formula(year_start: int, mode_rows: dict[str, int]) -> str:
    """Live Fleet-scale formula: Σ capacity_weight × fleet_size(mode) over the year
    block. Blank when no mode cell is filled."""
    year_col = _col(year_start + _YEAR_OFFSET)
    cells = [f"{year_col}{mode_rows[m]}" for m, _label in FLEET_MODES]
    guard = "COUNT(" + ",".join(cells) + ")=0"
    terms = "+".join(
        f"{MODE_CAPACITY_WEIGHT[m]}*N({year_col}{mode_rows[m]})" for m, _label in FLEET_MODES
    )
    return f'=IF({guard},"",{terms})'


# --- Export ------------------------------------------------------------------


def export_workbook(repo, path: str, years: list[int]) -> dict:
    """Build the per-agency time-series workbook and save it to `path`.

    One tab per agency (plus How-to + Data Dictionary). Pre-fills white cells from
    current DB values where present; the DB is usually empty at MVP. Returns a
    summary dict.
    """
    from openpyxl import Workbook

    from .dictionary import load_dictionary

    specs = load_dictionary()
    names = {c: s.display_name for c, s in specs.items()}
    meanings = {c: s.plain_meaning for c, s in specs.items()}
    index = _period_index(repo)

    wb = Workbook()
    _build_howto_sheet(wb)  # replaces the default sheet
    _build_dictionary_sheet(wb, names, meanings)

    filled = 0
    for slug, short_name in AGENCY_NAMES.items():
        filled += _build_agency_sheet(wb, slug, short_name, years, names, repo, index)

    wb.save(path)

    return {
        "path": path,
        "agencies": len(AGENCY_NAMES),
        "years": list(years),
        "metric_rows": len(METRIC_ROWS),
        "fleet_modes": len(FLEET_MODES),
        "filled_cells": filled,
    }


def _period_index(repo) -> dict:
    """Existing reporting periods keyed by (period_type, start, end) -> period."""
    return {
        (p.period_type, p.start_date, p.end_date): p
        for p in repo.list_reporting_periods()
    }


def _read_total_values(repo, agency_id, period_id, codes, mode_id=None) -> dict[str, object]:
    """Current 'total'-scope values for `codes` at a period, filtered to `mode_id`
    (None = system-wide; a mode's id for per-mode fleet)."""
    by_mid = {repo.metric_id(c): c for c in codes}
    out: dict[str, object] = {}
    for v in repo.list_current_values_for_agency_period(agency_id, period_id):
        if v.mode_id != mode_id or v.service_scope != "total":
            continue
        code = by_mid.get(v.metric_id)
        if code is not None:
            out[code] = v.value
    return out


def _build_howto_sheet(wb) -> None:
    """Sheet 1: plain-language instructions; reuses the default first sheet."""
    from openpyxl.styles import Alignment, Font

    ws = wb.active
    ws.title = SHEET_HOWTO
    ws.sheet_view.showGridLines = False

    lines: list[tuple[str, bool]] = [
        ("How to use this workbook", True),
        ("", False),
        ("This workbook collects transit numbers, one tab per agency. On an "
         "agency's tab, each row is a metric and the columns run left to right by "
         "calendar year. Within a year you see the twelve months, the four "
         "quarters, a year-to-date total, and the full-year total.", False),
        ("", False),
        ("What do I type?", True),
        ("- Ridership and Operating revenue: type each MONTH you have (the white "
         "month cells). The quarter, year-to-date, and full-year totals fill in "
         "automatically.", False),
        ("- Every other yearly number (service hours, costs, fleet age, and so "
         "on): type it once, in that row's YEAR column. The month and quarter "
         "cells for those rows are greyed out -- you don't use them.", False),
        ("- Fleet: at the bottom of each tab, type the number of vehicles for each "
         "mode (Bus, Subway, Light rail, Commuter rail, Streetcar) in the YEAR "
         "column. The 'Fleet scale' row is worked out for you.", False),
        ("", False),
        ("Leave a cell BLANK if you don't have the number yet -- never guess. The "
         "website keeps the last known number, so a blank here is fine.", False),
        ("", False),
        ("The colour code", True),
        ("- WHITE cells: type here.", False),
        ("- GREY cells: worked out automatically (quarter / year totals, ratios, "
         "Fleet scale). Don't type in them -- anything you put there is ignored "
         "and recalculated on the server.", False),
        ("", False),
        ("A note on years", True),
        ("Most agencies report on the calendar year. Metrolinx and BC Transit end "
         "their financial year in March; for them, the year column labelled e.g. "
         "'2024' means their 2024-25 fiscal year.", False),
        ("", False),
        ("When you're done", True),
        ("Save the file and run the import command. Your numbers go into the "
         "database; the grey totals and ratios are recomputed on the server.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=14 if (bold and i == 1) else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def _build_dictionary_sheet(wb, names: dict[str, str], meanings: dict[str, str]) -> None:
    """Sheet 2: one row per metric (all 32) -- meaning, unit, type, formula,
    native frequency."""
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(SHEET_DICT)
    headers = ["Column", "Plain meaning", "Unit", "Type", "Formula", "Native frequency"]
    for col, head in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=head).font = Font(bold=True)

    row = 2
    for code, meta in METRICS.items():
        # fleet_capacity is is_derived=False (sourced per-mode), but it IS computed
        # in the Fleet block, so show it as calculated for the reader.
        is_calculated = meta["is_derived"] or code == _FLEET_CAPACITY
        ws.cell(row=row, column=1, value=names[code])
        ws.cell(row=row, column=2, value=meanings[code])
        ws.cell(row=row, column=3, value=meta["unit"])
        ws.cell(row=row, column=4, value="Calculated" if is_calculated else "Sourced")
        ws.cell(row=row, column=5, value=meta["formula"] or ("Σ capacity_weight × fleet_size(mode)" if code == _FLEET_CAPACITY else ""))
        ws.cell(row=row, column=6, value=_native_frequency(code))
        row += 1

    ws.freeze_panes = "A2"
    for col_letter, width in (
        ("A", 28), ("B", 55), ("C", 10), ("D", 12), ("E", 38), ("F", 16)
    ):
        ws.column_dimensions[col_letter].width = width
    for r in range(2, row):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")


def _build_agency_sheet(wb, slug, short_name, years, names, repo, index) -> int:
    """One agency tab: a metric x (year-block) grid + a per-mode Fleet block.

    Returns the number of white cells pre-filled from the DB.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    from .periods import annual_period, monthly_period

    ws = wb.create_sheet(short_name)
    grey = PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
    bold = Font(bold=True)
    agency_id = repo.agency_id(slug)

    # --- header rows ---------------------------------------------------------
    ws.cell(row=_TITLE_ROW, column=_LABEL_COL, value=short_name).font = Font(bold=True, size=14)
    ws.cell(row=_SUBHEADER_ROW, column=_LABEL_COL, value="Metric").font = bold
    last_col = _FIRST_YEAR_COL + len(years) * YEAR_BLOCK_WIDTH - 1
    for yi, year in enumerate(years):
        ystart = _year_start_col(yi)
        yh = ws.cell(row=_YEAR_HEADER_ROW, column=ystart, value=int(year))
        yh.font = bold
        yh.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=_YEAR_HEADER_ROW, start_column=ystart,
            end_row=_YEAR_HEADER_ROW, end_column=ystart + YEAR_BLOCK_WIDTH - 1,
        )
        for off, (label, _kind) in enumerate(_YEAR_SUBHEADERS):
            c = ws.cell(row=_SUBHEADER_ROW, column=ystart + off, value=label)
            c.font = bold
            c.alignment = Alignment(horizontal="center")

    # Data-row index per metric (needed for the derived Year formulas).
    metric_row: dict[str, int] = {
        code: _FIRST_DATA_ROW + i for i, code in enumerate(METRIC_ROWS)
    }

    # --- metric rows ---------------------------------------------------------
    filled = 0
    for code in METRIC_ROWS:
        row = metric_row[code]
        is_monthly = code in MONTHLY_METRICS
        is_derived = METRICS[code]["is_derived"]
        ws.cell(row=row, column=_LABEL_COL, value=names[code]).font = bold
        fmt = _excel_unit_format(code)

        for yi, year in enumerate(years):
            ystart = _year_start_col(yi)
            month_cols = [ystart + off for off in _MONTH_OFFSETS]

            if is_monthly:
                # White month cells (pre-filled from monthly DB values).
                for mi, mcol in enumerate(month_cols, start=1):
                    cell = ws.cell(row=row, column=mcol)
                    cell.number_format = fmt
                    mp = monthly_period(year, mi)
                    period = index.get((mp.period_type, mp.start, mp.end))
                    if period is not None:
                        vals = _read_total_values(repo, agency_id, period.id, [code])
                        if code in vals:
                            cell.value = float(vals[code])
                            filled += 1
                # Grey Q (sum of its 3 months), YTD + Year (sum of all 12).
                for qi, qoff in enumerate(_QUARTER_OFFSETS):
                    qcell = ws.cell(row=row, column=ystart + qoff)
                    qcell.value = _sum_formula(row, month_cols[qi * 3:qi * 3 + 3])
                    qcell.number_format = fmt
                    qcell.fill = grey
                for off in (_YTD_OFFSET, _YEAR_OFFSET):
                    c = ws.cell(row=row, column=ystart + off)
                    c.value = _sum_formula(row, month_cols)
                    c.number_format = fmt
                    c.fill = grey
            else:
                # Annual-only: grey-lock months/quarters/YTD; the Year cell is the
                # only entry point (white for sourced, grey formula for derived).
                for off in range(YEAR_BLOCK_WIDTH):
                    if off == _YEAR_OFFSET:
                        continue
                    ws.cell(row=row, column=ystart + off).fill = grey
                ycell = ws.cell(row=row, column=ystart + _YEAR_OFFSET)
                ycell.number_format = fmt
                if is_derived:
                    ycell.value = _derived_year_formula(code, ystart, metric_row)
                    ycell.fill = grey
                else:
                    ap = annual_period(slug, year)
                    period = index.get((ap.period_type, ap.start, ap.end))
                    if period is not None:
                        vals = _read_total_values(repo, agency_id, period.id, [code])
                        if code in vals:
                            ycell.value = float(vals[code])
                            filled += 1

    # --- Fleet block ---------------------------------------------------------
    fleet_header_row = _FIRST_DATA_ROW + len(METRIC_ROWS) + 1
    ws.cell(row=fleet_header_row, column=_LABEL_COL, value="Fleet (vehicles by mode)").font = bold
    mode_row: dict[str, int] = {
        m: fleet_header_row + 1 + i for i, (m, _label) in enumerate(FLEET_MODES)
    }
    cap_row = fleet_header_row + 1 + len(FLEET_MODES)
    fmt_fleet = _excel_unit_format(_FLEET_SIZE)

    for m, label in FLEET_MODES:
        row = mode_row[m]
        ws.cell(row=row, column=_LABEL_COL, value=f"Fleet — {label}").font = bold
        mode_id = repo.mode_id(m)
        for yi, year in enumerate(years):
            ystart = _year_start_col(yi)
            # Grey-lock everything but the Year cell (fleet is point-in-time annual).
            for off in range(YEAR_BLOCK_WIDTH):
                if off == _YEAR_OFFSET:
                    continue
                ws.cell(row=row, column=ystart + off).fill = grey
            ycell = ws.cell(row=row, column=ystart + _YEAR_OFFSET)
            ycell.number_format = fmt_fleet
            ap = annual_period(slug, year)
            period = index.get((ap.period_type, ap.start, ap.end))
            if period is not None:
                vals = _read_total_values(repo, agency_id, period.id, [_FLEET_SIZE], mode_id=mode_id)
                if _FLEET_SIZE in vals:
                    ycell.value = float(vals[_FLEET_SIZE])
                    filled += 1

    # Fleet scale (fleet_capacity): grey computed cell per year (Year col only).
    ws.cell(row=cap_row, column=_LABEL_COL, value="Fleet scale").font = bold
    for yi, _year in enumerate(years):
        ystart = _year_start_col(yi)
        for off in range(YEAR_BLOCK_WIDTH):
            if off == _YEAR_OFFSET:
                continue
            ws.cell(row=cap_row, column=ystart + off).fill = grey
        c = ws.cell(row=cap_row, column=ystart + _YEAR_OFFSET)
        c.value = _fleet_capacity_formula(ystart, mode_row)
        c.number_format = _excel_unit_format(_FLEET_CAPACITY)
        c.fill = grey

    # --- presentation --------------------------------------------------------
    ws.freeze_panes = ws.cell(row=_FIRST_DATA_ROW, column=_FIRST_YEAR_COL)
    ws.column_dimensions["A"].width = 26
    for col in range(_FIRST_YEAR_COL, last_col + 1):
        ws.column_dimensions[_col(col)].width = 11
    return filled


# --- Import ------------------------------------------------------------------


def import_workbook(repo, path: str) -> dict:
    """Read the white cells of every agency tab and push them through the pipeline.

    For each agency tab: month cells -> monthly periods; annual sourced Year cells
    -> the agency's annual period; per-mode Fleet Year cells -> fleet_size at that
    mode_id. Then stage -> promote -> roll monthly ridership/revenue up to the
    year -> recompute derived ratios for every touched (agency, period) ->
    aggregate per-mode fleet into fleet_capacity -> refresh ranks. Grey cells
    (Q / YTD / Year roll-ups, derived ratios, Fleet scale) are never read; blank
    cells are skipped. Returns counts plus any warnings.
    """
    from decimal import Decimal, InvalidOperation

    from openpyxl import load_workbook

    from .contract import MetricValueRecord, SourceRef
    from .dictionary import load_dictionary
    from .jobs.derived_recompute import recompute_derived
    from .jobs.fleet_capacity_aggregate import fleet_capacity_aggregate
    from .jobs.rank_refresh import refresh_ranks
    from .jobs.rollup import calendar_rollup_metric, rollup_metric
    from .periods import annual_period, monthly_period
    from .promotion import promote_approved
    from .staging import stage_records

    names = {c: s.display_name for c, s in load_dictionary().items()}
    code_of_name = {name: code for code, name in names.items()}
    fleet_label_of = {f"Fleet — {label}": m for m, label in FLEET_MODES}
    source = SourceRef(
        document_type="manual_entry",
        extraction_method="manual",
        title="Manual entry (workbook import)",
    )
    wb = load_workbook(path, data_only=False)
    warnings: list[str] = []
    records: list[MetricValueRecord] = []
    monthly_agency_years: set[tuple[str, int]] = set()

    def as_decimal(cell, label, agency_name, period_label):
        try:
            return Decimal(str(cell))
        except (InvalidOperation, ValueError):
            warnings.append(
                f"non-numeric {label} for {agency_name} {period_label}: {cell!r}"
            )
            return None

    def add_record(slug, code, period, value, mode_code=None):
        meta = METRICS[code]
        records.append(
            MetricValueRecord(
                agency_slug=slug,
                metric_code=code,
                period_type=period.period_type,
                period_start=period.start,
                period_end=period.end,
                period_label=period.label,
                service_scope="total",
                value=str(value),
                unit=meta["unit"],
                quality="verified",
                mode_code=mode_code,
                currency="CAD" if meta["unit_type"] == "currency" else None,
                comparable_flag=code not in NON_RANKABLE_METRICS,
                source=source,
            )
        )

    for slug, short_name in AGENCY_NAMES.items():
        if short_name not in wb.sheetnames:
            continue
        ws = wb[short_name]
        # Map each year-block to (year, 1-based start column) from the header row.
        year_blocks: list[tuple[int, int]] = []
        col = _FIRST_YEAR_COL
        while col <= ws.max_column:
            raw = ws.cell(row=_YEAR_HEADER_ROW, column=col).value
            if raw is None:
                break
            try:
                year_blocks.append((int(raw), col))
            except (TypeError, ValueError):
                warnings.append(f"{short_name}: unreadable year header {raw!r}")
            col += YEAR_BLOCK_WIDTH

        for r in range(_FIRST_DATA_ROW, ws.max_row + 1):
            label = ws.cell(row=r, column=_LABEL_COL).value
            if not label:
                continue
            label = str(label).strip()

            # --- a per-mode Fleet row (white = Year cell, imports fleet_size) ---
            if label in fleet_label_of:
                mode_code = fleet_label_of[label]
                for year, ystart in year_blocks:
                    cell = ws.cell(row=r, column=ystart + _YEAR_OFFSET).value
                    if cell is None or cell == "":
                        continue
                    period = annual_period(slug, year)
                    value = as_decimal(cell, label, short_name, period.label)
                    if value is not None:
                        add_record(slug, _FLEET_SIZE, period, value, mode_code=mode_code)
                continue

            # --- a metric row -------------------------------------------------
            code = code_of_name.get(label)
            if code is None or code not in METRIC_ROWS:
                continue  # Fleet scale, separators, unknown rows: skip
            if METRICS[code]["is_derived"]:
                continue  # derived Year cell is a grey formula -> never imported

            is_monthly = code in MONTHLY_METRICS
            for year, ystart in year_blocks:
                if is_monthly:
                    for mi, off in enumerate(_MONTH_OFFSETS, start=1):
                        cell = ws.cell(row=r, column=ystart + off).value
                        if cell is None or cell == "":
                            continue
                        period = monthly_period(year, mi)
                        value = as_decimal(cell, label, short_name, period.label)
                        if value is not None:
                            add_record(slug, code, period, value)
                            monthly_agency_years.add((slug, year))
                else:
                    cell = ws.cell(row=r, column=ystart + _YEAR_OFFSET).value
                    if cell is None or cell == "":
                        continue
                    period = annual_period(slug, year)
                    value = as_decimal(cell, label, short_name, period.label)
                    if value is not None:
                        add_record(slug, code, period, value)

    # --- Orchestration: stage -> promote -> roll up -> recompute -> rank -----
    pending_ids = stage_records(repo, records, tier=0, feed_code="manual_entry")
    promoted = promote_approved(repo)

    periods: set[int] = set()
    agency_periods: set[tuple[str, int]] = set()
    for rec in records:
        pid = repo.get_or_create_reporting_period(
            rec.period_type, rec.period_start, rec.period_end, rec.period_label
        )
        periods.add(pid)
        agency_periods.add((rec.agency_slug, pid))

    # Roll monthly ridership + revenue up to the year BEFORE recompute, so annual
    # ratios (average_fare, ...) derive from the rolled-up annual inputs.
    rolled = 0
    for slug, year in sorted(monthly_agency_years):
        for code in MONTHLY_METRICS:
            written = rollup_metric(repo, slug, year, code)
            rolled += len(written.value_ids)
            for pid in written.period_ids:
                periods.add(pid)
                agency_periods.add((slug, pid))

    # Then fill the CALENDAR quarter / YTD / annual_calendar values from the same
    # monthly feeds (for the per-agency time-series grid). Runs AFTER the native
    # roll-up: calendar_rollup_metric writes only into EMPTY slots and cross-checks
    # existing ones, so it never clobbers a native annual -- it just adds the calendar
    # quarters / ytd (and, for fiscal-year agencies, the calendar-year total).
    for slug, year in sorted(monthly_agency_years):
        for code in MONTHLY_METRICS:
            cal = calendar_rollup_metric(repo, slug, year, code)
            rolled += len(cal.value_ids)
            warnings.extend(cal.warnings)
            for pid in cal.period_ids:
                periods.add(pid)
                agency_periods.add((slug, pid))

    derived = 0
    for agency_slug, pid in sorted(agency_periods):
        res = recompute_derived(repo, agency_slug, pid)
        derived += len(res.ids)
        warnings.extend(res.warnings)

    # Aggregate per-mode fleet sizes into fleet_capacity AFTER per-mode fleet
    # values exist (promoted above), and BEFORE rank refresh so the new aggregate
    # is ranked this run.
    for agency_slug, pid in sorted(agency_periods):
        fleet = fleet_capacity_aggregate(repo, agency_slug, pid)
        derived += len(fleet.value_ids)

    for pid in periods:
        for code in METRICS:
            refresh_ranks(repo, code, pid, service_scope="total")

    return {
        "staged": len(pending_ids),
        "promoted": len(promoted),
        "rolled": rolled,
        "derived": derived,
        "periods": len(periods),
        "warnings": warnings,
    }
